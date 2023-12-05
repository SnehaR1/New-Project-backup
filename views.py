from .models import ProductVariant  
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from userauth.models import CustomUser
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.db.models import Q
from django.utils import timezone
from django.db.models import Count ,F
from .models import Category,Brand, Product,ProductVariant,ProductImages,Color,Quantity,Order,OrderItem,Coupon,Offers,Blogs,BlogAdditionalImage
from adminapp.forms import UserEditForm,AddProduct,ProductVariantForm,ProductImagesForm,OrderItemForm,CouponForm,OffersForm,BlogForm,BlogImagesForm
from datetime import timedelta
from django.db.models import Sum
from django.db.models.functions import TruncWeek, TruncMonth, TruncYear
from django.http import JsonResponse
from django.db.models.functions import ExtractYear, ExtractMonth, ExtractWeek
import calendar
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font


def admin_login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        user = authenticate(request, username=email, password=password)
        
        if user is not None:
            if user.is_staff:
                login(request, user)
             
                return redirect('admin_dashboard')
            else:
                messages.error(request, 'This account does not have admin privileges.')
                return render(request, 'admin/admin_login.html')
        else:
            messages.error(request, 'Incorrect Username or Password')
            return render(request, 'admin/admin_login.html')
    return render(request, 'admin/admin_login.html')




@login_required(login_url='/admin/')
def admin_dashboard(request):
    filter_type="weekly"
    if request.method == 'POST':
        filter_type = request.POST.get('filter')
    orders=Order.objects.all()
    order_items=OrderItem.objects.all()
    products=Product.objects.all()
    total_revenue=Order.objects.aggregate(Sum('total_price')).get('total_price__sum', 0)
    total_products_sold=OrderItem.objects.aggregate(Sum('quantity')).get('quantity__sum', 0)
    total_customers=CustomUser.objects.aggregate(Count("email")).get('email__count', 0)
    product_orders = OrderItem.objects.annotate(product_title=F('product__title')).values("product_title").annotate(count=Count("id")).annotate(price=F('price__old_price')).annotate(amount=(F("count")*F("price"))).annotate(status=F("product__status")).values("product_title", "count","price","amount","status")
    products=Product.objects.all()
    product_labels=[]
    count=[]
    for product in product_orders:
        product_labels.append(product["product_title"])
        count.append(product["count"])
    monthly_orders = Order.objects.annotate(month=ExtractMonth("created_at")).values("month").annotate(count=Count("id")).values("month", "count")

    month_labels = []
    total_monthly_orders = []
    for o in monthly_orders:
        month_labels.append(calendar.month_name[o['month']])
        total_monthly_orders.append(o["count"])


    weekly_orders = Order.objects.annotate(week=ExtractWeek("created_at")).values("week").annotate(count=Count("id")).values("week", "count")

    week_labels = []
    total_weekly_orders = []
    for o in weekly_orders:
        week_labels.append(f"Week {o['week']}")
        total_weekly_orders.append(o["count"])

    yearly_orders = Order.objects.annotate(year=ExtractYear("created_at")).values("year").annotate(count=Count("id")).values("year", "count")

    year_labels = []
    total_yearly_orders = []
    for o in yearly_orders:
        year_labels.append(str(o['year']))
        total_yearly_orders.append(o["count"])
    total_price=Order.objects.aggregate(Sum('total_price'))
    print(total_price)
    return render(request, 'admin/admin_dashboard.html', {
        "product_orders":product_orders,
        "products":products,
        "orders":orders,
        "order_items":order_items,
        "total_products_sold":total_products_sold,
        "total_revenue":total_revenue,
        "total_customers":total_customers,
        "month_orders": monthly_orders,
        "month_labels": month_labels,
        "total_monthly_orders": total_monthly_orders,
        "week_orders": weekly_orders,
        "week_labels": week_labels,
        "total_weekly_orders": total_weekly_orders,
        "year_orders": yearly_orders,
        "year_labels": year_labels,
        "total_yearly_orders": total_yearly_orders,
        "filter_type":filter_type,
        "total_price":total_price,
        "product_labels":product_labels,
        "count":count
    })

@login_required(login_url='/admin/')
def admin_users(request):
    users = CustomUser.objects.all()

    if request.method == 'POST':
        form = UserEditForm(request.POST)
        if form.is_valid():
            user_id_to_block = form.cleaned_data.get('user_id_to_block')
            user_to_block = get_object_or_404(CustomUser, pk=user_id_to_block)
            user_to_block.is_active = not user_to_block.is_active
            user_to_block.save()
    else:
        form = UserEditForm()

    return render(request, "admin/admin_users.html", {'users': users, 'form': form})


@user_passes_test(lambda u: u.is_superuser)  
def dlt_user(request, user_id): 
    user= get_object_or_404(CustomUser, id=user_id)
    user.delete()
    messages.success(request, 'User deleted successfully!')
    return redirect('admin_users')

@login_required(login_url='/admin/')
def admin_category(request):
    if request.method == "POST":
        category_name = request.POST.get('category_name')
        if category_name:  
            Category.objects.create(title=category_name)
        return redirect('admin_category') 

    categories = Category.objects.annotate(product_count=Count('product_set'))
    return render(request, "admin/admin_category.html", {"categories":categories})



@user_passes_test(lambda u: u.is_superuser)
def dlt_category(request, cid):
    if request.method=="POST":
        print("Attempt to delete brand with ID:", cid)  
        category = get_object_or_404(Category,id=cid) 
        category.delete()
        messages.success(request, 'Category deleted successfully!')
        return redirect('admin_category')


@login_required(login_url='/admin/')
def admin_brand(request):
    if request.method == "POST":
        brand_name = request.POST.get('brand_name')
        if brand_name:  
            Brand.objects.create(title=brand_name)
        return redirect('admin_brand')  

    brands = Brand.objects.annotate(product_count=Count('product_set'))
    return render(request, "admin/admin_brand.html", {"brands":brands})


def dlt_brand(request, bid): 
    if request.method=="POST":
        print("Attempt to delete brand with ID:", bid)  
        brand = get_object_or_404(Brand, id=bid)
        brand.delete()
        messages.success(request, 'Brand deleted successfully!')
        return redirect('admin_brand')

@login_required(login_url='/admin/')
def admin_products(request):
    products = Product.objects.all()
    return render(request, "admin/admin_products.html", {"products": products})


@login_required(login_url='/admin/')
def add_product(request, pid=None):
    if pid:
        product = get_object_or_404(Product, pk=pid)
        operation = "Edit"
    else:
        product = Product()
        operation = "Add"

    if request.method == 'POST':
        form = AddProduct(request.POST, request.FILES, instance=product)

        if form.is_valid():
            form.save()

            messages.success(request, f'Product {operation}ed successfully!')
            return redirect('admin_products')
    else:
        form = AddProduct(instance=product)

    context = {
        "form": form,
        "product": product
    }

    return render(request, "admin/add_product.html", context)

@login_required(login_url='/admin/')
def admin_variant(request, pid=None):
    products = Product.objects.all()
 
    context = {
        "products": products,
       
    }
    return render(request, "admin/admin_variants.html", context)

@login_required(login_url='/admin/')
def add_variant(request):
    if request.method == 'POST':
        form = ProductVariantForm(request.POST, request.FILES)
        img_form = ProductImagesForm(request.POST, request.FILES)

        if form.is_valid() and img_form.is_valid():
            product_variant = form.save(commit=False)

            new_color_name = request.POST.get('new_color')
            if new_color_name:
                color, created = Color.objects.get_or_create(name=new_color_name)
                product_variant.color = color

            new_quantity_name = request.POST.get('new_quantity')
            if new_quantity_name:
                quantity, created = Quantity.objects.get_or_create(name=new_quantity_name)
                product_variant.quantity = quantity

            product_variant.save()

        
            images = request.FILES.getlist('images')

            for image in images:
                ProductImages.objects.create(images=image, productvariant=product_variant)

            messages.success(request, 'Product and images added successfully!')
            return redirect('admin_variant')
        else:
            messages.error(request, 'Something went wrong')
    else:
        form = ProductVariantForm()
        img_form = ProductImagesForm()

    context = {
        'form': form,
        'img_form': img_form,
    }
    return render(request, 'admin/add_variant.html', context)

@user_passes_test(lambda u: u.is_superuser)  
def dlt_product(request, product_id): 
    product = get_object_or_404(Product, id=product_id)
    product.delete()
    messages.success(request, 'Product deleted successfully!')
    return redirect('admin_products')

@user_passes_test(lambda u: u.is_superuser)  
def dlt_variant(request, pv_id): 
    product_variant = get_object_or_404(ProductVariant, id=pv_id)
    product_variant.delete()
    messages.success(request, 'Product Variant deleted successfully!')
    return redirect('admin_variant')

@login_required(login_url='/admin/')
def logout_view(request):
    logout(request)
    messages.success(request, 'Logged out successfully!')
    return redirect('admin_login')


def edit_user(request, user_id):
    user = CustomUser.objects.get(pk=user_id)
    if request.method == "POST":
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            user = form.save(commit=False)  
            user.save()
            return redirect('admin_users')
    else:
        return redirect('admin_users')
    
@login_required(login_url='/admin/')
def admin_orders(request):
    orders = Order.objects.all()
    orderitems = OrderItem.objects.all()
    productvariants=Product.objects.select_related('productvariant').all()
    return render(request, 'admin/admin_orders.html', {"orders": orders, "productvairants":productvariants,"orderitems":orderitems})


def update_status(request, order_item_id):
    order_item = get_object_or_404(OrderItem, id=order_item_id)
    print(f"Processing order item with ID: {order_item_id}")
    if request.method == 'POST':
        print(f"Received POST request for order item ID {order_item_id}. Data: {request.POST}")
    
        order_item_form = OrderItemForm(request.POST, instance=order_item)
        if order_item_form.is_valid():
            new_status = order_item_form.cleaned_data['status']     
            if new_status == "delivered":
                order_item.delivery_date = timezone.now()  
        
            order_item.status = new_status
            try:
                order_item.save()
             
                print(f"Order item {order_item.id} saved successfully with status {new_status}")
            except Exception as e:
             
                print(f"Error saving order item {order_item.id}: {e}")
        else:
            print(f"Form for order item {order_item.id} is not valid. Errors: {order_item_form.errors}")
            messages.error(request, f"Form for order item {order_item.id} is not valid. Errors: {order_item_form.errors}")

    else:
       
        order_item_form = OrderItemForm(instance=order_item)
    context = {'order_item_form': order_item_form, 'order_item': order_item}

    return render(request, 'admin/admin_orders.html', context)





def toggle_listing(request, pv_id):
    product_variant = get_object_or_404(ProductVariant, id=pv_id)

    if request.method == 'POST':
        is_listed = request.POST.get('listing_status') == 'listed'
        product_variant.is_listed = is_listed
        product_variant.save()

    return render(request, 'admin/admin_variants.html', {'product_variant': product_variant})

def unlist_category(request,category_id):
    category=get_object_or_404(Category,id=category_id)
    if request.method=="POST":
        is_listed=request.POST.get('listing_status')=='listed'
        category.is_listed=is_listed
        category.save()
        return render(request,'admin/admin_category.html',{'category':category})

def unlist_brand(request,brand_id):
    brand=get_object_or_404(Brand,id=brand_id)
    if request.method=="POST":
        is_listed=request.POST.get('listing_status')=='listed'
        brand.is_listed=is_listed
        brand.save()
        return render(request,'admin/admin_brand.html',{'brand':brand})
    
@login_required(login_url='/admin/')   
def product_listing(request,product_id):
    product=get_object_or_404(Product,id=product_id)
    if request.method=="POST":
        is_listed=request.POST.get('listing_status')=='listed'
        product.is_listed=is_listed
        product.save()
        return render(request,'admin/admin_brand.html',{'product':product})

@login_required(login_url='/admin/')
def edit_variant(request, pv_id):
    product_variant = get_object_or_404(ProductVariant, id=pv_id)

    if request.method == 'POST':
        form = ProductVariantForm(request.POST, request.FILES, instance=product_variant)
        img_form = ProductImagesForm(request.POST, request.FILES, instance=product_variant)

        if form.is_valid() and img_form.is_valid():
            product_variant = form.save(commit=False)

            new_color_name = request.POST.get('new_color')
            if new_color_name:
                color, created = Color.objects.get_or_create(name=new_color_name)
                product_variant.color = color

            new_quantity_name = request.POST.get('new_quantity')
            if new_quantity_name:
                quantity, created = Quantity.objects.get_or_create(name=new_quantity_name)
                product_variant.quantity = quantity

            product_variant.save()

            images = request.FILES.getlist('images')
            if images:
                ProductImages.objects.filter(productvariant=product_variant).delete()

            
            for image in images:
                ProductImages.objects.create(images=image, productvariant=product_variant)

            messages.success(request, 'Product variant updated successfully!')
            return redirect('admin_variant')
        else:
            messages.error(request, 'Something went wrong')
    else:
        form = ProductVariantForm(instance=product_variant)
        img_form = ProductImagesForm(instance=product_variant)

    context = {
        'form': form,
        'img_form': img_form,
        'product_variant': product_variant,
    }

    return render(request, 'admin/edit_variant.html', context)

@login_required(login_url='/admin/')
def admin_coupon(request):
    coupons=Coupon.objects.all()
    return render(request,'admin/admin_coupon.html',{'coupons':coupons})

@login_required(login_url='/admin/')
def add_coupon(request):
    if request.method == "POST":
        form = CouponForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,"Coupon added Successfully!!!")
            return redirect('admin_coupon')
        else:
            error_message = "Failed to add the Coupon! Please correct the following errors: {}".format(form.errors)
            messages.error(request, error_message)
    else:
        form = CouponForm()

    return render(request, 'admin/add_coupon.html', {"form": form})

@login_required(login_url='/admin/')
def edit_coupon(request,c_id):
    coupon=get_object_or_404(Coupon, id=c_id)
    if request.method == "POST":
        form = CouponForm(request.POST, instance=coupon)
        if form.is_valid():
            updated_coupon = form.save(commit=False)
            updated_coupon.save()
            messages.success(request,"Coupon edited Successfully!!!")
            return redirect('admin_coupon')
        else:
            error_message = "Failed to edit the Coupon! Please correct the following errors: {}".format(form.errors)
            messages.error(request, error_message)
    else:
        form = CouponForm(instance=coupon)

    return render(request, 'admin/edit_coupon.html', {"form": form})

def admin_offers(request):
    offers=Offers.objects.all()
    return render(request,'admin/admin_offers.html',{'offers':offers})

def add_offers(request):
    if request.method=="POST":
        form=OffersForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,"Offer Added Successfully")
            return redirect('admin_offers')
        else:
            error_message = "Failed to add the Offer! Please correct the following errors: {}".format(form.errors)
            messages.error(request, error_message)
    else:
        form=OffersForm()
    return render(request,'admin/add_offer.html',{'form':form})

def report(request):
    orders=Order.objects.all()
    order_items=OrderItem.objects.all()
    products=Product.objects.all()
    total_revenue=Order.objects.aggregate(Sum('total_price')).get('total_price__sum', 0)
    total_products_sold=OrderItem.objects.aggregate(Sum('quantity')).get('quantity__sum', 0)
    total_customers=CustomUser.objects.aggregate(Count("email")).get('email__count', 0)
    product_orders = OrderItem.objects.annotate(product_title=F('product__title')).values("product_title").annotate(count=Count("id")).annotate(price=F('price__price')).annotate(amount=(F("count")*F("price"))).annotate(status=F("product__status")).values("product_title", "count","price","amount","status")
    products=Product.objects.all()
    product_labels=[]
    count=[]
    for product in product_orders:
        product_labels.append(product["product_title"])
        count.append(product["count"])
    monthly_orders = Order.objects.annotate(month=ExtractMonth("created_at")).values("month").annotate(count=Count("id")).values("month", "count")

    month_labels = []
    total_monthly_orders = []
    for o in monthly_orders:
        month_labels.append(calendar.month_name[o['month']])
        total_monthly_orders.append(o["count"])

    weekly_orders = Order.objects.annotate(week=ExtractWeek("created_at")).values("week").annotate(count=Count("id")).values("week", "count")

    week_labels = []
    total_weekly_orders = []
    for o in weekly_orders:
        week_labels.append(f"Week {o['week']}")
        total_weekly_orders.append(o["count"])

    yearly_orders = Order.objects.annotate(year=ExtractYear("created_at")).values("year").annotate(count=Count("id")).values("year", "count")

    year_labels = []
    total_yearly_orders = []
    for o in yearly_orders:
        year_labels.append(str(o['year']))
        total_yearly_orders.append(o["count"])
    total_price=Order.objects.aggregate(Sum('total_price'))
    template = get_template('admin/admin_dashboard_report.html')
    if request.GET.get('excel'):
        product_orders_for_excel = [
    {
        'product_title': item.get('product_title', ''),
        'count': item.get('count', ''),
        'price': item.get('price', ''),
        'amount': item.get('amount', ''),
        'status': 'In stock' if item.get('status', False) else 'Out of stock',
    }
    for item in product_orders
]
        workbook = openpyxl.Workbook()
        sheet = workbook.active

 
        headers = ['Product name', 'Sales', 'Unit Price', 'Amount', 'Status']
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)


        for row_num, item in enumerate(product_orders_for_excel, 2):
            sheet.cell(row=row_num, column=1, value=item['product_title'])
            sheet.cell(row=row_num, column=2, value=item['count'])
            sheet.cell(row=row_num, column=3, value=item['price'])
            sheet.cell(row=row_num, column=4, value=item['amount'])
            sheet.cell(row=row_num, column=5, value=item['status'])


        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=product_orders.xlsx'

    
        workbook.save(response)

        return response

    elif request.GET.get('pdf'):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'

        context={
        "product_orders":product_orders,
        "products":products,
        "orders":orders,
        "order_items":order_items,
        "total_products_sold":total_products_sold,
        "total_revenue":total_revenue,
        "total_customers":total_customers,
        "month_orders": monthly_orders,
        "month_labels": month_labels,
        "total_monthly_orders": total_monthly_orders,
        "week_orders": weekly_orders,
        "week_labels": week_labels,
        "total_weekly_orders": total_weekly_orders,
        "year_orders": yearly_orders,
        "year_labels": year_labels,
        "total_yearly_orders": total_yearly_orders,
        "total_price":total_price,
        "product_labels":product_labels,
        "count":count
    }
        template_content = template.render(context)

     
        pdf_data = pisa.CreatePDF(template_content, dest=response)
        
        if pdf_data.err:
            return HttpResponse('We had some errors <pre>' + template_content + '</pre>')
     
        return response

    
    
def edit_offers(request,offer_id):
    offer=get_object_or_404(Offers, id=offer_id)
    if request.method=="POST":
        form=OffersForm(request.POST, instance=offer)
        if form.is_valid():
            updated_offer = form.save(commit=False)
            updated_offer.save()
            messages.success(request,"Offer edited Successfully!!!")
            return redirect("admin_offers")
        else:
            error_message = "Failed to edit the Coupon! Please correct the following errors: {}".format(form.errors)
            messages.error(request, error_message)
    else:
        form=OffersForm(instance=offer)

    return render(request,"admin/edit_offers.html",{"form":form})

def admin_blog(request):
    blogs=Blogs.objects.all()
    return render(request,"admin/admin_blog.html",{"blogs":blogs})

def add_blog(request):
    if request.method == "POST":
        form = BlogForm(request.POST, request.FILES)
        img_form = BlogImagesForm(request.POST, request.FILES)
        if form.is_valid() and img_form.is_valid():
            blog_instance = form.save()
            additional_image_instance = img_form.save(commit=False)
            additional_image_instance.blog = blog_instance
            additional_image_instance.save()

            messages.success(request, "Blog Added Successfully!!!")
            return redirect("admin_blog")
        else:
            print(form.errors)
            messages.error(request, 'Please correct the errors below.')
    else:
        form = BlogForm()
        img_form = BlogImagesForm()

    return render(request, "admin/add_blog.html", {"form": form, "img_form": img_form})

@login_required(login_url='/admin/')
def edit_blog(request, blog_id):
    blog = get_object_or_404(Blogs, id=blog_id)

    if request.method == "POST":
        form = BlogForm(request.POST, request.FILES, instance=blog)
        if form.is_valid():
            try:
              
                blog = form.save()

            
                add_images = request.FILES.getlist('add_images')

                if add_images:
                    BlogAdditionalImage.objects.filter(blog=blog).delete()

                for image in add_images:
                    BlogAdditionalImage.objects.create(add_images=image, blog=blog)

                messages.success(request, 'Blog updated successfully!')
                return redirect('admin_blog')
            except Exception as e:
                print(f"Error: {e}")
                messages.error(request, 'Something went wrong')
        else:
            print(form.errors)
            messages.error(request, 'Please correct the errors below.')

    else:
        form = BlogForm(instance=blog)

    context = {
        'form': form,
        'blog': blog,
    }
    return render(request, "admin/edit_blog.html", context)


